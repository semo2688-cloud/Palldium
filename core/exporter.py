import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import config


def _to_dataframe(records: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(records, columns=config.COLUMNS)
    for col in config.COLUMNS:
        if col not in df.columns:
            df[col] = None
    # Keep integer columns as Int64 (nullable int) so None rows don't become float
    for col in ("연식", "가격"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
    return df[config.COLUMNS]


def to_csv_bytes(records: list[dict]) -> bytes:
    df = _to_dataframe(records)
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def to_excel_bytes(records: list[dict]) -> bytes:
    df = _to_dataframe(records)
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="중고에어컨")
        ws = writer.sheets["중고에어컨"]

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(config.COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            lengths = [len(str(col_name))]
            for i in range(min(len(df), 100)):
                val = df.iloc[i, col_idx - 1]
                if val is not None and str(val) != "None":
                    lengths.append(len(str(val)))
            ws.column_dimensions[col_letter].width = min(max(lengths) + 4, 50)

    return buf.getvalue()
